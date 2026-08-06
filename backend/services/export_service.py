"""
Reports & Export System (Module 12)
Generates native Excel (.xlsx) workbooks for:
  - Insider Threat Reports
  - Behavioral Analytics Reports
  - Investigation Reports
  - Compliance Reports
  - Risk Assessment Reports
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, desc
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Dict, Any

from backend.models.dataset import Employee, BehavioralAnomaly, Incident, EmployeeBaseline, EmployeeRiskHistory


class ExportService:
    @classmethod
    async def generate_excel_report(cls, db: AsyncSession, report_type: str) -> io.BytesIO:
        """
        Generate an Excel workbook (.xlsx) stream for the specified report category.
        """
        wb = Workbook()
        ws = wb.active

        # Style definitions
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")  # Indigo-600
        title_font = Font(name="Calibri", size=14, bold=True, color="0F172A")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        report_type_clean = report_type.lower().replace(" ", "_")

        if report_type_clean == "insider_threat":
            ws.title = "Insider Threat Report"
            ws.append(["ITBIS - Insider Threat Security Summary"])
            ws.append([])

            headers = ["Employee ID", "Full Name", "Department", "Role", "Risk Score", "Risk Category", "Active Anomalies"]
            ws.append(headers)

            stmt = select(Employee).options(selectinload(Employee.anomalies)).order_by(desc(Employee.risk_score))
            employees = (await db.execute(stmt)).scalars().all()
            for emp in employees:
                anom_count = len(emp.anomalies) if emp.anomalies else 0
                cat = "Critical Risk" if emp.risk_score >= 85 else "High Risk" if emp.risk_score >= 60 else "Medium Risk" if emp.risk_score >= 30 else "Low Risk"
                ws.append([emp.employee_id, emp.full_name, emp.department or "N/A", emp.role or "N/A", emp.risk_score, cat, anom_count])

        elif report_type_clean == "behavioral_analytics":
            ws.title = "Behavioral Analytics"
            ws.append(["ITBIS - Employee Behavioral Baselines"])
            ws.append([])

            headers = ["Employee ID", "Avg Daily Logons", "After-Hours Logon %", "Avg Daily USB", "Avg Daily Files", "Avg Emails Sent", "Cloud Upload %"]
            ws.append(headers)

            stmt = select(EmployeeBaseline)
            baselines = (await db.execute(stmt)).scalars().all()
            for b in baselines:
                ws.append([
                    b.employee_id,
                    round(b.avg_daily_logons, 2),
                    f"{round(b.after_hours_logon_ratio * 100, 1)}%",
                    round(b.avg_daily_usb_connects, 2),
                    round(b.avg_daily_file_accesses, 2),
                    round(b.avg_daily_emails_sent, 2),
                    f"{round(b.cloud_upload_ratio * 100, 1)}%"
                ])

        elif report_type_clean == "investigation":
            ws.title = "Investigation Cases"
            ws.append(["ITBIS - Threat Investigation Cases"])
            ws.append([])

            headers = ["Case ID", "Employee ID", "Title", "Severity", "Status", "Assigned Analyst", "Created Date"]
            ws.append(headers)

            stmt = select(Incident).order_by(desc(Incident.created_at))
            incidents = (await db.execute(stmt)).scalars().all()
            for inc in incidents:
                ws.append([
                    inc.incident_number,
                    inc.employee_id,
                    inc.title,
                    inc.severity,
                    inc.status,
                    inc.assigned_analyst or "Unassigned",
                    inc.created_at.strftime("%Y-%m-%d %H:%M")
                ])

        elif report_type_clean == "compliance":
            ws.title = "Compliance Audit"
            ws.append(["ITBIS - Platform Compliance & Access Audit"])
            ws.append([])

            headers = ["Employee ID", "Full Name", "Department", "System Access Status", "Policy Risk Classification"]
            ws.append(headers)

            stmt = select(Employee).order_by(Employee.employee_id)
            employees = (await db.execute(stmt)).scalars().all()
            for emp in employees:
                status = "Active" if emp.is_active else "Suspended"
                policy = "Compliant" if emp.risk_score < 60 else "Requires Security Review"
                ws.append([emp.employee_id, emp.full_name, emp.department or "General", status, policy])

        else:  # risk_assessment
            ws.title = "Risk Assessment"
            ws.append(["ITBIS - Comprehensive Risk Assessment Metrics"])
            ws.append([])

            headers = ["Employee ID", "Composite Risk Score", "Behavioral Score (35%)", "Privilege Score (25%)", "Data Access Score (20%)", "Access Pattern (10%)", "Historical Events (10%)"]
            ws.append(headers)

            stmt = select(EmployeeRiskHistory).order_by(desc(EmployeeRiskHistory.timestamp)).limit(200)
            histories = (await db.execute(stmt)).scalars().all()
            for h in histories:
                ws.append([
                    h.employee_id,
                    h.risk_score,
                    round(h.behavioral_score, 1),
                    round(h.privilege_score, 1),
                    round(h.data_access_score, 1),
                    round(h.access_pattern_score, 1),
                    round(h.historical_events_score, 1)
                ])

        # Style Title Row
        ws.cell(row=1, column=1).font = title_font

        # Style Table Header Row (Row 3)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style Data Cells & Auto-fit Column Widths
        for row in range(4, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
