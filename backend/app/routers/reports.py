
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill
import io
from ..database import get_db
from ..dependencies import get_current_user
from ..models import User, UserProfile
from ..services.risk_score import risk_score_calculator
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)
@router.get("/")
def generate_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    total_users = db.query(User).count()
    active_users = db.query(User).filter(User.is_active == True).count()
    high_risk_users = db.query(UserProfile).filter(UserProfile.risk_score >= 50).count()

    report = {
        "total_users": total_users,
        "active_users": active_users,
        "high_risk_users": high_risk_users,
        "generated_by": "Insider Threat Behavioral Intelligence System",
    }

    return report

@router.get("/export/risk-assessment/excel")
def export_risk_assessment_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    profiles = db.query(UserProfile).order_by(UserProfile.risk_score.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risk Assessment"

    headers = ["Employee ID", "Department", "Designation", "Risk Score", "Risk Category"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for profile in profiles:
        category = risk_score_calculator.categorize(profile.risk_score)
        ws.append([
            profile.employee_id,
            profile.department,
            profile.designation,
            profile.risk_score,
            category,
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=risk_assessment_report.xlsx"},
    )

@router.get("/export/investigations/excel")
def export_investigations_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from ..models import Incident
    incidents = db.query(Incident).order_by(Incident.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Investigations"

    headers = ["Incident ID", "Employee ID", "Risk Category", "Risk Score", "Status", "Assigned Analyst", "Created At"]
    ws.append(headers)

    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for i in incidents:
        ws.append([
            i.id,
            i.employee_id,
            i.risk_category,
            i.risk_score_at_creation,
            i.status,
            i.assigned_analyst or "Unassigned",
            i.created_at.strftime("%Y-%m-%d %H:%M") if i.created_at else "",
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=investigation_report.xlsx"},
    )

@router.get("/export/summary/pdf")
def export_summary_pdf(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from ..models import Incident, Alert
    from sqlalchemy import func as sqlfunc

    total_employees = db.query(UserProfile).count()
    avg_risk = db.query(sqlfunc.avg(UserProfile.risk_score)).scalar() or 0
    high_risk_count = db.query(UserProfile).filter(UserProfile.risk_score >= 50).count()
    total_incidents = db.query(Incident).count()
    open_incidents = db.query(Incident).filter(Incident.status.in_(["Open", "Investigating"])).count()
    total_alerts = db.query(Alert).count()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("TitleStyle", parent=styles["Heading1"], fontSize=18, spaceAfter=20)
    elements.append(Paragraph("Insider Threat Behavioral Intelligence System", title_style))
    elements.append(Paragraph("Insider Threat Risk Assessment Report", styles["Heading2"]))
    elements.append(Spacer(1, 20))

    data = [
        ["Metric", "Value"],
        ["Total Employees Monitored", str(total_employees)],
        ["Organizational Avg Risk Score", str(round(avg_risk, 2))],
        ["High/Critical Risk Employees", str(high_risk_count)],
        ["Total Incidents", str(total_incidents)],
        ["Open Incidents", str(open_incidents)],
        ["Total Alerts", str(total_alerts)],
    ]

    table = Table(data, colWidths=[300, 150])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=insider_threat_summary_report.pdf"},
    )